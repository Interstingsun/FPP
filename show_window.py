from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from img.MainWindow import Ui_MainWindow       #导入主窗体文件中的ui类
import sys                      # 导入系统模块
import openpyxl
import numpy as np
import matplotlib.ticker as mticker
import matplotlib.pyplot as plt
import time

#主窗体初始化类
class Main(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super(Main,self).__init__()
        self.setupUi(self)
        self.total_dic = {}
        self.day_sort = {}
        self.max_wue = {}
        self.plants_dic = {}
        self.gen_dic = {}
        self.gen_sort = {}
        self.board = {}

    def click_find_file_path(self):
        # 设置文件扩展名过滤，同一个类型的不同格式如xlsx和xls 用空格隔开
        filename, filetype = QFileDialog.getOpenFileName(self, "选取Excel文件", "/Users/mengmeng/Desktop/",
                                                         "Excel Files (*.xls *.xlsx)")
        self.load_xls(filename)

    def click_find_plants_path(self):
        # 设置文件扩展名过滤，同一个类型的不同格式如xlsx和xls 用空格隔开
        filename, filetype = QFileDialog.getOpenFileName(self, "选取Excel文件", "/Users/mengmeng/Desktop/",
                                                         "Excel Files (*.xls *.xlsx)")
        self.load_plants(filename)
    
    def load_plants(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb['Plants']
        max_row = ws.max_row
        max_column = ws.max_column
        for column in range(1, max_column + 1):
            if ws.cell(1, column).value == 'PlantName':
                plantname_column = column
            if ws.cell(1, column).value == '#A-genotype':
                gen_column = column
        gen_dic = {}
        for row in range(2, max_row + 1):
            item1 = ws.cell(row, plantname_column)
            item2 = float(ws.cell(row, gen_column).value)
            gen_dic.update({item1.value: item2})
        self.gen_dic = dict(sorted(gen_dic.items(), key=lambda item: item[1]))
        gen_sort = 1
        for item in self.gen_dic:
            self.gen_sort.update({item: gen_sort})
            gen_sort = gen_sort + 1
        dialog = QDialog()
        button = QPushButton('确定', dialog)
        button.clicked.connect(dialog.close)
        button.move(50, 50)
        dialog.setWindowTitle('加载完成')
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.exec()

    def click_smooth_curve(self):
        # 设置文件扩展名过滤，同一个类型的不同格式如xlsx和xls 用空格隔开
        filename, filetype = QFileDialog.getOpenFileName(self, "选取Excel文件", "/Users/mengmeng/Desktop/",
                                                         "Excel Files (*.xls *.xlsx)")
        self.load_xls(filename)

    def sort_dict(self, time_value):
        time_sort = {}
        dict_no = 1
        for item in time_value:
            time_sort.update({dict_no: item})
            dict_no = dict_no + 1
        return time_sort

    def search_row(self, sheets, row=1, column=1):
        while sheets.cell(row, column).value:
            row = row + 1
        return row

    def search_column(self, sheets, row=1, column=1):
        while sheets.cell(row, column).value:
            column = column + 1
        return column

    def smooth(self, x, M):
        x = np.array(list(x))
        K = round(M / 2 - 0.1)
        lenx = len(x)
        if lenx < 2 * K + 1:
            print("数据长度小于平滑点数")
        else:
            y = np.zeros(lenx)
            for NN in range(0, lenx, 1):
                startind = max(0, NN - K)
                endind = min(NN + K + 1, lenx)
                y[NN] = np.mean(x[startind:endind])
        return y

    def load_xls(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb['WUE']
        max_row = ws.max_row
        max_column = ws.max_column
        datetime_column = 0
        for column in range(1, max_column + 1):
            if ws.cell(2, column).is_date:
                datetime_column = column
        for column in range(1, max_column + 1):
            if column == datetime_column:
                continue
            elif 'vpd' in ws.cell(1, column).value.lower():
                continue
            else:
                day_no = 0
                for row in range(2, max_row + 1):
                    item = ws.cell(row, column)
                    day = str(ws.cell(row, datetime_column).value).split(' ')[0]
                    if not day in self.day_sort:
                        day_no = day_no + 1
                        self.day_sort.update({day: day_no})
                    try:
                        time = str(ws.cell(row, datetime_column).value).split(' ')[1]
                    except:
                        break
                    if item.data_type == 'n' and item.value is not None:
                        try:
                            tmpDic = self.total_dic[str(ws.cell(1, column).value) + '@' + day]
                        except:
                            self.total_dic.update({str(ws.cell(1, column).value) + '@' + day: 0})
                            tmpDic = {}
                        tmpDic.update({time: item.value})
                    else:
                        try:
                            tmpDic = self.total_dic[str(ws.cell(1, column).value) + '@' + day]
                        except:
                            self.total_dic.update({str(ws.cell(1, column).value) + '@' + day: 0})
                            tmpDic = {}
                        tmpDic.update({time: 0})
                    # plants+date
                    self.total_dic[str(ws.cell(1, column).value) + '@' + day] = tmpDic
        dialog = QDialog()
        button = QPushButton('确定', dialog)
        button.clicked.connect(dialog.close)
        button.move(50, 50)
        dialog.setWindowTitle('加载完成')
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.exec()

    def get_broken_line(self):
        if not self.total_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载WUE表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.plants_dic:
            for plants_day in self.total_dic:
                plants = plants_day.split('@')[0]
                day = plants_day.split('@')[1]
                origin_dic = self.total_dic[plants_day]
                for item in origin_dic:
                    try:
                        tmpDic = self.plants_dic[plants]
                    except:
                        tmpDic = {}
                        self.plants_dic.update({plants: tmpDic})
                    tmpDic.update({day + ' ' + item: origin_dic[item]})
                    self.plants_dic.update({plants: tmpDic})
        plt.figure(figsize=(20, 10), dpi=50)
        for plants in self.plants_dic:
            time = list(self.plants_dic[plants].keys())
            wue = list(self.plants_dic[plants].values())
            plt.plot(time, wue, label=plants)
            day = '0'
            tmp_list = []
            total = 0
            for i in range(len(time)):
                if time[i].split(' ')[0] != day:
                    try:
                        y1_max = np.argmax(tmp_list)
                        plt.plot(y1_max + total, wue[y1_max + total], 'ko')
                    except:
                        y1_max = 0
                    total = total + len(tmp_list)
                    tmp_list = []
                    day = time[i].split(' ')[0]
                    tmp_list.append(wue[i])
                else:
                    tmp_list.append(wue[i])
            y1_max = np.argmax(tmp_list)
            plt.plot(y1_max + total, wue[y1_max + total], 'ko')

        plt.title("WUE broken line")
        plt.legend(loc=9, ncol=6)
        plt.gca().xaxis.set_major_locator(mticker.MultipleLocator(960))
        plt.show()

    def get_curve(self):
        if not self.total_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载WUE表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.plants_dic:
            for plants_day in self.total_dic:
                plants = plants_day.split('@')[0]
                day = plants_day.split('@')[1]
                origin_dic = self.total_dic[plants_day]
                for item in origin_dic:
                    try:
                        tmpDic = self.plants_dic[plants]
                    except:
                        tmpDic = {}
                        self.plants_dic.update({plants: tmpDic})
                    tmpDic.update({day + ' ' + item: origin_dic[item]})
                    self.plants_dic.update({plants: tmpDic})
        plt.figure(figsize=(20, 10), dpi=50)
        for plants in self.plants_dic:
            time = list(self.plants_dic[plants].keys())
            wue = list(self.plants_dic[plants].values())
            new_wue = self.smooth(wue, 30)
            plt.plot(time, new_wue, label=plants)
            day = '0'
            tmp_list = []
            total = 0
            for i in range(len(time)):
                if time[i].split(' ')[0] != day:
                    try:
                        y1_max = np.argmax(tmp_list)
                        plt.plot(y1_max + total, new_wue[y1_max + total], 'ko')
                    except:
                        y1_max = 0
                    total = total + len(tmp_list)
                    tmp_list = []
                    day = time[i].split(' ')[0]
                    tmp_list.append(new_wue[i])
                else:
                    tmp_list.append(new_wue[i])
            y1_max = np.argmax(tmp_list)
            plt.plot(y1_max + total, new_wue[y1_max + total], 'ko')

        plt.title("WUE curve")
        plt.legend(loc=9, ncol=6)
        plt.gca().xaxis.set_major_locator(mticker.MultipleLocator(960))
        plt.show()

    def get_maxwue(self):
        if not self.total_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载WUE表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.gen_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载基因表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        for plants_day in self.total_dic:
            origin_dic = self.total_dic[plants_day]
            max_value = sorted(origin_dic.items(), key=lambda item: item[1], reverse=True)[0]
            self.max_wue.update({plants_day: max_value})
        new_wb = openpyxl.Workbook()
        ws = new_wb['Sheet']
        ws.cell(1, 1, value="单元")
        ws.cell(2, 1, value="genotype")
        for item in self.max_wue:
            # item={B2@2022-10-01}
            # max_wue[item]=('07:57:00', 0.287350002033926)
            plants = item.split('@')[0]
            genotype = self.gen_dic[plants]
            wue = self.max_wue[item]
            day = item.split('@')[1]
            day_sort = self.day_sort[day]
            gen_sort = self.gen_sort[plants]
            offset = 1
            ws.cell(row=offset, column=1, value="max")
            ws.cell(row=1 + offset, column=1, value="单元")
            ws.cell(row=2 + offset, column=1, value="genotype")
            column_letter = openpyxl.utils.get_column_letter(1)
            ws.column_dimensions[column_letter].width = 30
            ws.cell(row=1 + offset, column=gen_sort + 1, value=plants)
            ws.cell(row=2 + offset, column=gen_sort + 1, value=genotype)
            ws.cell(row=day_sort + 2 + offset, column=1, value=day)
            column_letter = openpyxl.utils.get_column_letter(gen_sort + 1)
            ws.column_dimensions[column_letter].width = 20
            ws.cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=wue[0][0:5])
            offset = len(self.day_sort) + 5
            ws.cell(row=offset, column=1, value="max_value")
            ws.cell(row=1 + offset, column=1, value="单元")
            ws.cell(row=2 + offset, column=1, value="genotype")
            ws.cell(row=1 + offset, column=gen_sort + 1, value=plants)
            ws.cell(row=2 + offset, column=gen_sort + 1, value=genotype)
            ws.cell(row=day_sort + 2 + offset, column=1, value=day)
            ws.cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=wue[1])
        new_wb.save('max_wue.xlsx')
        dialog = QDialog()
        button = QPushButton('确定', dialog)
        button.clicked.connect(dialog.close)
        button.move(50, 50)
        dialog.setWindowTitle('表格已保存！')
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.exec()

    def half_to_max(self):
        if not self.max_wue:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未计算WUE最大值！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.gen_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载基因表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        new_wb = openpyxl.Workbook()
        for item in self.max_wue:
            # item={B2@2022-10-01}
            origin_value = self.max_wue[item]
            # 表格以日期为sheet的名称索引
            day = item.split('@')[1]
            half = origin_value[1] / 2
            origin_dic = self.total_dic[item]
            try:
                s_column = self.gen_sort[item.split('@')[0]]
                column_letter = openpyxl.utils.get_column_letter(s_column)
                new_wb[day].column_dimensions[column_letter].width = 45
                new_wb[day].cell(row=1, column=s_column, value=item.split('@')[0])
                new_wb[day].cell(row=2, column=s_column, value=self.gen_dic[item.split('@')[0]])
                new_wb[day].cell(row=3, column=s_column, value=str(half) + '-' + str(origin_value[1]))
            except KeyError as e:
                new_wb.create_sheet(day)
                s_column = self.gen_sort[item.split('@')[0]]
                column_letter = openpyxl.utils.get_column_letter(s_column)
                new_wb[day].column_dimensions[column_letter].width = 45
                new_wb[day].cell(row=1, column=s_column, value=item.split('@')[0])
                new_wb[day].cell(row=2, column=s_column, value=self.gen_dic[item.split('@')[0]])
                new_wb[day].cell(row=3, column=s_column, value=str(half) + '-' + str(origin_value[1]))
            for time in origin_dic:
                plants = item.split('@')[0]
                s_row = self.search_row(new_wb[day], 1, self.gen_sort[plants])
                if origin_dic[time] >= half:
                    new_wb[day].cell(s_row, self.gen_sort[plants], value=item.split('@')[1] + ' ' + time[0:5])
        moren = new_wb['Sheet']
        new_wb.remove(moren)
        new_wb.save('half_to_max_wue.xlsx')
        self.twfive_to_max()

    def twfive_to_max(self):
        if not self.max_wue:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未计算WUE最大值！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.gen_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载基因表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        new_wb = openpyxl.Workbook()
        for item in self.max_wue:
            # item={B2@2022-10-01}
            origin_value = self.max_wue[item]
            day = item.split('@')[1]
            half = origin_value[1] / 4
            origin_dic = self.total_dic[item]
            try:
                s_column = self.gen_sort[item.split('@')[0]]
                column_letter = openpyxl.utils.get_column_letter(s_column)
                new_wb[day].column_dimensions[column_letter].width = 45
                new_wb[day].cell(row=1, column=s_column, value=item.split('@')[0])
                new_wb[day].cell(row=2, column=s_column, value=self.gen_dic[item.split('@')[0]])
                new_wb[day].cell(row=3, column=s_column, value=str(half) + '-' + str(origin_value[1]))
            except KeyError as e:
                new_ws = new_wb.create_sheet(day)
                s_column = self.gen_sort[item.split('@')[0]]
                column_letter = openpyxl.utils.get_column_letter(s_column)
                new_wb[day].column_dimensions[column_letter].width = 45
                new_wb[day].cell(row=1, column=s_column, value=item.split('@')[0])
                new_wb[day].cell(row=2, column=s_column, value=self.gen_dic[item.split('@')[0]])
                new_wb[day].cell(row=3, column=s_column, value=str(half) + '-' + str(origin_value[1]))
            for time in origin_dic:
                plants = item.split('@')[0]
                s_row = self.search_row(new_wb[day], 1, self.gen_sort[plants])
                if origin_dic[time] >= half:
                    new_wb[day].cell(s_row, self.gen_sort[plants], value=item.split('@')[1] + ' ' + time)
        moren = new_wb['Sheet']
        new_wb.remove(moren)
        new_wb.save('25_to_max_wue.xlsx')
        self.sefive_to_max()

    def sefive_to_max(self):
        if not self.max_wue:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未计算WUE最大值！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.gen_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载基因表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        new_wb = openpyxl.Workbook()
        for item in self.max_wue:
            # item={B2@2022-10-01}
            origin_value = self.max_wue[item]
            day = item.split('@')[1]
            half = origin_value[1] - origin_value[1] / 4
            origin_dic = self.total_dic[item]
            try:
                s_column = self.gen_sort[item.split('@')[0]]
                column_letter = openpyxl.utils.get_column_letter(s_column)
                new_wb[day].column_dimensions[column_letter].width = 45
                new_wb[day].cell(row=1, column=s_column, value=item.split('@')[0])
                new_wb[day].cell(row=2, column=s_column, value=self.gen_dic[item.split('@')[0]])
                new_wb[day].cell(row=3, column=s_column, value=str(half) + '-' + str(origin_value[1]))
            except KeyError as e:
                new_ws = new_wb.create_sheet(day)
                s_column = self.gen_sort[item.split('@')[0]]
                column_letter = openpyxl.utils.get_column_letter(s_column)
                new_wb[day].column_dimensions[column_letter].width = 45
                new_wb[day].cell(row=1, column=s_column, value=item.split('@')[0])
                new_wb[day].cell(row=2, column=s_column, value=self.gen_dic[item.split('@')[0]])
                new_wb[day].cell(row=3, column=s_column, value=str(half) + '-' + str(origin_value[1]))
            for time in origin_dic:
                plants = item.split('@')[0]
                s_row = self.search_row(new_wb[day], 1, self.gen_sort[plants])
                if origin_dic[time] >= half:
                    new_wb[day].cell(s_row, self.gen_sort[plants], value=item.split('@')[1] + ' ' + time)
        moren = new_wb['Sheet']
        new_wb.remove(moren)
        new_wb.save('75_to_max_wue.xlsx')
        dialog = QDialog()
        button = QPushButton('确定', dialog)
        button.clicked.connect(dialog.close)
        button.move(50, 50)
        dialog.setWindowTitle('表格已保存！')
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.exec()

    def find_range(self):
        def time_interval(left, right):
            lhh = int(left.split(':')[0])
            lmm = int(left.split(':')[1])
            rhh = int(right.split(':')[0])
            rmm = int(right.split(':')[1])
            mm = rmm - lmm if rmm - lmm >= 0 else rmm - lmm + 60
            hh = rhh - lhh if rmm - lmm >= 0 else rhh - lhh - 1
            return str(hh).zfill(2)+':'+str(mm).zfill(2)

        if not self.max_wue:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未计算WUE最大值！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        new_wb = openpyxl.Workbook()
        for i in range(3):
            center_type = str(i * 25 + 25)
            for item in self.max_wue:
                wue = self.max_wue[item]
                center_time = wue[0]
                center_value = wue[1]
                if center_type == '25':
                    board_value = center_value / 4
                elif center_type == '50':
                    board_value = center_value / 2
                else:
                    board_value = center_value - center_value / 4
                end = False
                board_right = center_time
                board_left = center_time
                while not end:
                    board_right = self.search_board(item, board_value, board_right, 'right')
                    end, board_right = self.is_end(item, board_value, board_right, 'right')
                end = False
                while not end:
                    board_left = self.search_board(item, board_value, board_left, 'left')
                    end, board_left = self.is_end(item, board_value, board_left, 'left')
                self.board.update({item: center_time + '-' + board_left + '-' + board_right})
            new_wb.create_sheet(center_type)
            for board_item in self.board:
                day = board_item.split('@')[1]
                plants = board_item.split('@')[0]
                gen_sort = self.gen_sort[plants]
                day_sort = self.day_sort[day]
                genotype = self.gen_dic[plants]
                center_time = self.board[board_item].split('-')[0]
                board_left = self.board[board_item].split('-')[1]
                board_right = self.board[board_item].split('-')[2]
                offset = 1
                new_wb[center_type].cell(row=offset, column=1, value="board_left")
                new_wb[center_type].cell(row=1 + offset, column=1, value="单元")
                new_wb[center_type].cell(row=2 + offset, column=1, value="genotype")
                column_letter = openpyxl.utils.get_column_letter(1)
                new_wb[center_type].column_dimensions[column_letter].width = 30
                new_wb[center_type].cell(row=1 + offset, column=gen_sort + 1, value=plants)
                new_wb[center_type].cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=1, value=day)
                column_letter = openpyxl.utils.get_column_letter(gen_sort + 1)
                new_wb[center_type].column_dimensions[column_letter].width = 20
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=board_left[0:5])
                offset = len(self.day_sort) + 5
                new_wb[center_type].cell(row=offset, column=1, value="board_right")
                new_wb[center_type].cell(row=1 + offset, column=1, value="单元")
                new_wb[center_type].cell(row=2 + offset, column=1, value="genotype")
                new_wb[center_type].cell(row=1 + offset, column=gen_sort + 1, value=plants)
                new_wb[center_type].cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=1, value=day)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=board_right[0:5])
                offset = 2 * (len(self.day_sort) + 4) + 1
                new_wb[center_type].cell(row=offset, column=1, value="left_duration")
                new_wb[center_type].cell(row=1 + offset, column=1, value="单元")
                new_wb[center_type].cell(row=2 + offset, column=1, value="genotype")
                new_wb[center_type].cell(row=1 + offset, column=gen_sort + 1, value=plants)
                new_wb[center_type].cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=1, value=day)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=gen_sort + 1,
                                         value=time_interval(board_left, center_time))
                offset = 3 * (len(self.day_sort) + 4) + 1
                new_wb[center_type].cell(row=offset, column=1, value="right_duration")
                new_wb[center_type].cell(row=1 + offset, column=1, value="单元")
                new_wb[center_type].cell(row=2 + offset, column=1, value="genotype")
                new_wb[center_type].cell(row=1 + offset, column=gen_sort + 1, value=plants)
                new_wb[center_type].cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=1, value=day)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=gen_sort + 1,
                                         value=time_interval(center_time, board_right))
                offset = 4 * (len(self.day_sort) + 4) + 1
                new_wb[center_type].cell(row=offset, column=1, value="duration")
                new_wb[center_type].cell(row=1 + offset, column=1, value="单元")
                new_wb[center_type].cell(row=2 + offset, column=1, value="genotype")
                new_wb[center_type].cell(row=1 + offset, column=gen_sort + 1, value=plants)
                new_wb[center_type].cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=1, value=day)
                new_wb[center_type].cell(row=day_sort + 2 + offset, column=gen_sort + 1,
                                         value=time_interval(board_left, board_right))
        moren = new_wb['Sheet']
        new_wb.remove(moren)
        new_wb.save('range.xlsx')
        dialog = QDialog()
        button = QPushButton('确定', dialog)
        button.clicked.connect(dialog.close)
        button.move(50, 50)
        dialog.setWindowTitle('表格已保存！')
        dialog.setWindowModality(Qt.ApplicationModal)
        dialog.exec()

    def search_board(self, item, board_value, board, direction='left'):
        origin_dic = self.total_dic[item]
        time_sort = self.sort_dict(origin_dic)
        for key, value in time_sort.items():
            if board == value:
                center_sort = key
                break
        if direction == 'left':
            next_board = center_sort - 1
        else:
            next_board = center_sort + 1
        current = center_sort
        if next_board not in time_sort:
            return time_sort[current]
        while origin_dic[time_sort[next_board]] >= board_value:
            current = next_board
            if direction == 'left':
                next_board = next_board - 1
            else:
                next_board = next_board + 1
            if next_board not in time_sort:
                return time_sort[current]
        return time_sort[current]

    def is_end(self, item, board_value, board, direction='left'):
        origin_dic = self.total_dic[item]
        time_sort = self.sort_dict(origin_dic)
        for key, value in time_sort.items():
            if board == value:
                end_time = key
                break
        if direction == 'left':
            end1 = end_time - 2
            end2 = end_time - 3
        else:
            end1 = end_time + 2
            end2 = end_time + 3
        if end1 in time_sort and origin_dic[time_sort[end1]] < board_value:
            if end2 in time_sort and origin_dic[time_sort[end2]] >= board_value:
                is_end = False
                return is_end, time_sort[end2]
            else:
                is_end = True
                return is_end, time_sort[end_time]
        elif end1 in time_sort and origin_dic[time_sort[end1]] >= board_value:
            is_end = False
            return is_end, time_sort[end1]
        else:
            is_end = True
            return is_end, time_sort[end_time]

    def percent(self):
        if not self.total_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载WUE表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        if not self.gen_dic:
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('未加载基因表格！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()
            return
        range_wue_dic={}
        total_wue_dic={}
        range_percent_dic = {}
        start, sok = QInputDialog.getText(self, '生长比例', 'wue起点')
        if sok:
            end, ok = QInputDialog.getText(self, '生长比例', 'wue终点')
        if ok:
            if ':' in start:
                start_hh = int(start.split(':')[0])
                start_mm = int(start.split(':')[1])
            else:
                start_hh = int(start)
                start_mm = 0
            if ':' in end:
                end_hh = int(end.split(':')[0])
                end_mm = int(end.split(':')[1])
            else:
                end_hh = int(end)
                end_mm = 0
            start_hm = start_hh * 60 + start_mm
            end_hm = end_hh * 60 + end_mm
            for plants_day in self.total_dic:
                origin_dic = self.total_dic[plants_day]
                total_wue = 0
                range_wue = 0
                for time in origin_dic:
                    time_tmp = int(time[0:2]) * 60 + int(time[3:5])
                    if start_hm <= time_tmp <= end_hm:
                        range_wue = range_wue + origin_dic[time]
                    total_wue = total_wue + origin_dic[time]
                    tmp_percent = 0
                    if total_wue != 0:
                        tmp_percent = range_wue/total_wue
                range_wue_dic.update({plants_day: range_wue})
                total_wue_dic.update({plants_day: total_wue})
                range_percent_dic.update({plants_day: tmp_percent})
            new_wb = openpyxl.Workbook()
            ws = new_wb['Sheet']
            ws.cell(1, 1, value="单元")
            ws.cell(2, 1, value="genotype")
            for item in range_percent_dic:
                # item={B2@2022-10-01}
                # max_wue[item]=('07:57:00', 0.287350002033926)
                plants = item.split('@')[0]
                genotype = self.gen_dic[plants]
                range_wue = range_wue_dic[item]
                total_wue = total_wue_dic[item]
                tmp_percent = range_percent_dic[item]
                day = item.split('@')[1]
                day_sort = self.day_sort[day]
                gen_sort = self.gen_sort[plants]
                offset = 1
                ws.cell(row=offset, column=1, value="range wue")
                ws.cell(row=1 + offset, column=1, value="单元")
                ws.cell(row=2 + offset, column=1, value="genotype")
                column_letter = openpyxl.utils.get_column_letter(1)
                ws.column_dimensions[column_letter].width = 30
                ws.cell(row=1 + offset, column=gen_sort + 1, value=plants)
                ws.cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                ws.cell(row=day_sort + 2 + offset, column=1, value=day)
                column_letter = openpyxl.utils.get_column_letter(gen_sort + 1)
                ws.column_dimensions[column_letter].width = 20
                ws.cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=range_wue)
                offset = len(self.day_sort) + 5
                ws.cell(row=offset, column=1, value="total wue")
                ws.cell(row=1 + offset, column=1, value="单元")
                ws.cell(row=2 + offset, column=1, value="genotype")
                column_letter = openpyxl.utils.get_column_letter(1)
                ws.column_dimensions[column_letter].width = 30
                ws.cell(row=1 + offset, column=gen_sort + 1, value=plants)
                ws.cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                ws.cell(row=day_sort + 2 + offset, column=1, value=day)
                column_letter = openpyxl.utils.get_column_letter(gen_sort + 1)
                ws.column_dimensions[column_letter].width = 20
                ws.cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=total_wue)
                offset = 2 * (len(self.day_sort) + 4) + 1
                ws.cell(row=offset, column=1, value="range precent")
                ws.cell(row=1 + offset, column=1, value="单元")
                ws.cell(row=2 + offset, column=1, value="genotype")
                column_letter = openpyxl.utils.get_column_letter(1)
                ws.column_dimensions[column_letter].width = 30
                ws.cell(row=1 + offset, column=gen_sort + 1, value=plants)
                ws.cell(row=2 + offset, column=gen_sort + 1, value=genotype)
                ws.cell(row=day_sort + 2 + offset, column=1, value=day)
                column_letter = openpyxl.utils.get_column_letter(gen_sort + 1)
                ws.column_dimensions[column_letter].width = 20
                ws.cell(row=day_sort + 2 + offset, column=gen_sort + 1, value=tmp_percent)

            new_wb.save(start.replace(':', '')+'_'+end.replace(':', '')+'.xlsx')
            dialog = QDialog()
            button = QPushButton('确定', dialog)
            button.clicked.connect(dialog.close)
            button.move(50, 50)
            dialog.setWindowTitle('表格已保存！')
            dialog.setWindowModality(Qt.ApplicationModal)
            dialog.exec()

if __name__=="__main__":
    app=QApplication(sys.argv)
    #主窗体对象
    main=Main()
    #加载表格，按钮事件
    main.btn_1.triggered.connect(main.click_find_file_path)
    # 上传基因顺序表格，并重新排序按钮事件
    main.btn_2.triggered.connect(main.click_find_plants_path)
    # 查找最大值并生成表格，按钮事件
    main.btn_3.triggered.connect(main.get_maxwue)
    # 查找范围值并生成表格，按钮事件
    main.btn_4.triggered.connect(main.half_to_max)
    # 显示折线图，按钮事件
    main.btn_5.triggered.connect(main.get_broken_line)
    # 平滑曲线，按钮事件
    main.btn_6.triggered.connect(main.get_curve)
    # 寻找时间段
    main.btn_7.triggered.connect(main.find_range)
    # 计算百分比
    main.btn_8.triggered.connect(main.percent)
    #显示主窗体
    main.show()
    sys.exit(app.exec_())

